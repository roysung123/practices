#------------selenium 用於模擬瀏覽器---------------------------
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
#--------------------------------------------------------------
#------------beautifulsoup 用於解析網頁原始碼------------------
from bs4 import BeautifulSoup
#--------------------------------------------------------------
#---------------------雜項-------------------------------------
import re
import time
import json
import codecs
from openpyxl import Workbook
#--------------------------------------------------------------

#--------code start--------------------------------------------
#----先下載chrome driver並放至指定位址-------------------------
driver = webdriver.Chrome('./chromedriver')#指定在同一個資料夾下
driver.get("https://wsts.most.gov.tw/STSWeb/Award/AwardMultiQuery.aspx") #前往這個網址

#-------------模擬使用者操作到我們要的頁面---------------------
#                                   btnAward : "獎勵"按鈕
#                          dtlItem_btnItem_0 : "傑出研究獎"按鈕
#                          dtlItem_btnItem_2 : "吳大猷先生紀念獎"按鈕
# wUctlAwardQueryPage$repQuery$ctl01$ddlYRst : "年份"第一個下拉式選單
#   wUctlAwardQueryPage$repQuery$ctl02$ddlO1 : "單位類型"下拉式選單
#   wUctlAwardQueryPage$repQuery$ctl02$ddlO2 : "單位名稱"下拉式選單
#            wUctlAwardQueryPage$ddlPageSize : "每頁筆數"下拉式選單
#               wUctlAwardQueryPage$btnQuery : "查詢"按鈕

awardbtn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.NAME, "btnAward")))#獎勵#等待直到可點擊or超過十秒
awardbtn.click()

gsabtn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "dtlItem_btnItem_0")))#傑出研究獎
gsabtn.click()

yearselect = Select(driver.find_element_by_name('wUctlAwardQueryPage$repQuery$ctl01$ddlYRst'))#年份
yearselect.select_by_index(len(yearselect.options)-1)#選至最舊的年份

pgselect = Select(driver.find_element_by_name('wUctlAwardQueryPage$repQuery$ctl02$ddlO1'))#單位類型
pgselect.select_by_visible_text(u"公立大學")

time.sleep(1)

pselect = Select(driver.find_element_by_name('wUctlAwardQueryPage$repQuery$ctl02$ddlO2'))#單位名稱
pselect.select_by_visible_text(u"國立臺灣大學")

pagesize = Select(driver.find_element_by_name('wUctlAwardQueryPage$ddlPageSize'))#每頁筆數
pagesize.select_by_visible_text(u"200")

sereach = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.NAME, "wUctlAwardQueryPage$btnQuery")))#查詢
sereach.click()
#--------------------------------------------------------------

#----------到達頁面 截下原始碼 關閉模擬器----------------------
soup = BeautifulSoup(driver.page_source, "lxml")
driver.quit()
#--------------------------------------------------------------

#--------------------存入excel---------------------------------
finalList = []
wb = Workbook()
ws = wb.active

ws['E1'] = '處室'
ws['D1'] = '計畫名稱'
ws['C1'] = '任職機關'
ws['B1'] = '獲獎人姓名'
ws['A1'] = '獲獎年度'

for table in soup.select('#wUctlAwardQueryPage_grdResult'):
    for trs in table.find_all(class_=re.compile("Grid_Row|Grid_AlternatingRow")):
        str = ""
        for i in range(0,len(trs.select('td'))):
            finalList.append(trs.select('td')[i].text)
        ws.append(finalList)
        finalList.clear()

wb.save('Awardss.xlsx')
#--------------------------------------------------------------
